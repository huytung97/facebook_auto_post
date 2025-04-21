from openpyxl import load_workbook

SCRIPT_WB = 'Script'
GROUP_WB = 'Group'
PROFILE_WB = 'Profile'

def load_sheet(wb, sheet_name, id_col) -> dict:
    result = dict()

    current_wb = wb[sheet_name]
    headers = [cell.value for cell in next(current_wb.iter_rows(min_row=1, max_row=1))]
    for row in current_wb.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))

        id = row_dict.pop(id_col)
        result[id] = row_dict
    
    return result

def load_data(excel_file_path) -> tuple:
    wb = load_workbook(excel_file_path)

    scripts = load_sheet(wb, SCRIPT_WB, 'id')

    groups = load_sheet(wb, GROUP_WB, 'group_url')
    for group_id, data in groups.items():
        data['members'] = data['members'].split('\n')
        groups[group_id] = data

    profiles = load_sheet(wb, PROFILE_WB, 'name')

    return scripts, groups, profiles